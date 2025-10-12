"""
Excel File Parser Utilities

Provides Excel file parsing functionality using pandas and openpyxl.
Handles OEWS-specific data formats including special values ('#', '*', NaN).
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import pandas as pd
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Excel file parsing utilities for OEWS data migration

    Handles reading Excel files, extracting sheet data, and parsing
    OEWS-specific formats with special values and data patterns.
    """

    def __init__(self):
        """Initialize the Excel parser"""
        self.special_values = {'#', '*'}  # OEWS suppressed data markers

    def parse_file(self, file_path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
        """
        Parse an Excel file and return all sheets as DataFrames

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary mapping sheet names to DataFrames

        Raises:
            FileNotFoundError: If the Excel file doesn't exist
            ValueError: If the file cannot be parsed
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        logger.info(f"Parsing Excel file: {file_path}")

        try:
            # Read all sheets from the Excel file
            sheets_dict = pd.read_excel(
                file_path,
                sheet_name=None,  # Read all sheets
                engine='openpyxl'
            )

            logger.info(f"Successfully parsed {len(sheets_dict)} sheets from {file_path.name}")
            return sheets_dict

        except Exception as e:
            logger.error(f"Failed to parse Excel file {file_path}: {str(e)}")
            raise ValueError(f"Cannot parse Excel file: {str(e)}")

    def parse_with_types(self, file_path: Union[str, Path]) -> List[Dict[str, Any]]:
        """
        Parse Excel file and convert data with type inference

        Args:
            file_path: Path to the Excel file

        Returns:
            List of dictionaries representing rows with inferred types
        """
        sheets = self.parse_file(file_path)

        result = []
        for sheet_name, df in sheets.items():
            # Convert DataFrame to list of dictionaries
            records = df.to_dict('records')

            # Add sheet metadata
            for record in records:
                record['_sheet_name'] = sheet_name

            result.extend(records)

        logger.info(f"Parsed {len(result)} total records from {file_path}")
        return result

    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """
        Get list of sheet names from an Excel file

        Args:
            file_path: Path to the Excel file

        Returns:
            List of sheet names
        """
        file_path = Path(file_path)

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()

            logger.debug(f"Found {len(sheet_names)} sheets in {file_path.name}")
            return sheet_names

        except Exception as e:
            logger.error(f"Failed to read sheet names from {file_path}: {str(e)}")
            raise ValueError(f"Cannot read Excel file metadata: {str(e)}")

    def get_sheet_data(
        self,
        file_path: Union[str, Path],
        sheet_name: str,
        skip_rows: int = 0,
        max_rows: Optional[int] = None
    ) -> pd.DataFrame:
        """
        Get data from a specific sheet

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read
            skip_rows: Number of rows to skip from the top
            max_rows: Maximum number of rows to read

        Returns:
            DataFrame containing the sheet data
        """
        file_path = Path(file_path)

        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                skiprows=skip_rows,
                nrows=max_rows,
                engine='openpyxl'
            )

            logger.debug(f"Read {len(df)} rows from sheet '{sheet_name}' in {file_path.name}")
            return df

        except Exception as e:
            logger.error(f"Failed to read sheet '{sheet_name}' from {file_path}: {str(e)}")
            raise ValueError(f"Cannot read sheet data: {str(e)}")

    def detect_header_row(self, file_path: Union[str, Path], sheet_name: str) -> int:
        """
        Detect which row contains the header in a sheet

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet

        Returns:
            Row index (0-based) where headers are found
        """
        file_path = Path(file_path)

        try:
            # Read first few rows to detect headers
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                nrows=10,
                header=None,
                engine='openpyxl'
            )

            # Look for row with mostly non-null string values
            for idx, row in df.iterrows():
                non_null_count = row.notna().sum()
                string_count = row.apply(lambda x: isinstance(x, str)).sum()

                # Header row typically has many non-null string values
                if non_null_count > len(row) * 0.7 and string_count > len(row) * 0.5:
                    logger.debug(f"Detected header row at index {idx} in sheet '{sheet_name}'")
                    return int(idx)

            # Default to first row if no clear header found
            logger.debug(f"No clear header row found in sheet '{sheet_name}', using row 0")
            return 0

        except Exception as e:
            logger.warning(f"Failed to detect header row in {file_path}: {str(e)}")
            return 0

    def get_column_names(self, file_path: Union[str, Path], sheet_name: str) -> List[str]:
        """
        Get column names from a sheet

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet

        Returns:
            List of column names
        """
        header_row = self.detect_header_row(file_path, sheet_name)

        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            skiprows=header_row,
            nrows=0,
            engine='openpyxl'
        )

        columns = df.columns.tolist()
        logger.debug(f"Found {len(columns)} columns in sheet '{sheet_name}'")
        return columns

    def is_oews_special_value(self, value: Any) -> bool:
        """
        Check if a value is an OEWS special value ('#', '*', etc.)

        Args:
            value: Value to check

        Returns:
            True if the value is a special OEWS marker
        """
        if pd.isna(value):
            return False

        value_str = str(value).strip()
        return value_str in self.special_values

    def get_file_metadata(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Extract metadata from an Excel file

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary containing file metadata
        """
        file_path = Path(file_path)

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            metadata = {
                'file_name': file_path.name,
                'file_path': str(file_path.absolute()),
                'file_size_bytes': file_path.stat().st_size,
                'file_size_mb': round(file_path.stat().st_size / (1024 * 1024), 2),
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'properties': {
                    'creator': workbook.properties.creator,
                    'title': workbook.properties.title,
                    'subject': workbook.properties.subject,
                    'created': workbook.properties.created,
                    'modified': workbook.properties.modified,
                }
            }

            workbook.close()

            logger.debug(f"Extracted metadata from {file_path.name}")
            return metadata

        except Exception as e:
            logger.error(f"Failed to extract metadata from {file_path}: {str(e)}")
            return {
                'file_name': file_path.name,
                'file_path': str(file_path.absolute()),
                'error': str(e)
            }

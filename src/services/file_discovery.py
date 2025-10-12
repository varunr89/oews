"""
File Discovery Service Implementation

Implements FileDiscoveryService contract for discovering Excel files.
Maps to FR-001: Excel file discovery and enumeration.
"""

import hashlib
import logging
import time
from pathlib import Path
from typing import Optional, List
from datetime import datetime
import re
import os

from openpyxl import load_workbook

from src.contracts.file_discovery_service import (
    FileDiscoveryService,
    ExcelFileInfo,
    DiscoveryOptions,
    DiscoveryResult
)

logger = logging.getLogger(__name__)


class FileDiscoveryServiceImpl(FileDiscoveryService):
    """
    Implementation of FileDiscoveryService for Excel file discovery

    Discovers, enumerates, and validates Excel files in directory structures
    with support for filtering, exclusion patterns, and metadata extraction.
    """

    def __init__(self):
        """Initialize the file discovery service"""
        self.logger = logging.getLogger(__name__)

    def discover_excel_files(
        self,
        directory_path: Path,
        options: Optional[DiscoveryOptions] = None
    ) -> DiscoveryResult:
        """
        Discover all Excel files in the specified directory

        Args:
            directory_path: Root directory to search
            options: Discovery configuration options

        Returns:
            DiscoveryResult containing found files and metadata

        Raises:
            ValueError: If directory_path is invalid
            PermissionError: If directory is not accessible
        """
        start_time = time.time()

        # Validate directory
        if not directory_path or not isinstance(directory_path, Path):
            raise ValueError("directory_path must be a valid Path object")

        if not directory_path.exists():
            raise ValueError(f"Directory does not exist: {directory_path}")

        if not directory_path.is_dir():
            raise ValueError(f"Path is not a directory: {directory_path}")

        # Check directory accessibility
        if not os.access(directory_path, os.R_OK):
            raise PermissionError(f"Directory is not readable: {directory_path}")

        # Use default options if none provided
        if options is None:
            options = DiscoveryOptions()

        files_found: List[ExcelFileInfo] = []
        errors: List[str] = []
        total_size = 0

        # Compile exclude patterns
        exclude_regexes = [re.compile(pattern) for pattern in options.exclude_patterns]

        try:
            # Determine search pattern
            if options.include_subdirectories:
                search_patterns = [f"**/*{ext}" for ext in options.file_extensions]
            else:
                search_patterns = [f"*{ext}" for ext in options.file_extensions]

            # Find all Excel files
            for pattern in search_patterns:
                for file_path in directory_path.glob(pattern):
                    try:
                        # Skip if not a file
                        if not file_path.is_file():
                            continue

                        # Check exclude patterns
                        if any(regex.match(file_path.name) for regex in exclude_regexes):
                            self.logger.debug(f"Excluding file: {file_path.name}")
                            continue

                        # Check file size
                        file_size = file_path.stat().st_size
                        if file_size > options.max_file_size:
                            self.logger.warning(
                                f"Skipping file (exceeds max size): {file_path.name} ({file_size} bytes)"
                            )
                            errors.append(
                                f"File exceeds max size limit: {file_path.name}"
                            )
                            continue

                        # Get file metadata
                        try:
                            file_info = self.get_file_metadata(file_path)
                            files_found.append(file_info)
                            total_size += file_size

                            self.logger.debug(f"Discovered file: {file_path.name}")

                        except Exception as e:
                            error_msg = f"Failed to get metadata for {file_path.name}: {str(e)}"
                            self.logger.error(error_msg)
                            errors.append(error_msg)

                    except Exception as e:
                        error_msg = f"Error processing {file_path}: {str(e)}"
                        self.logger.error(error_msg)
                        errors.append(error_msg)

        except Exception as e:
            error_msg = f"Error during file discovery: {str(e)}"
            self.logger.error(error_msg)
            errors.append(error_msg)

        duration = time.time() - start_time

        self.logger.info(
            f"Discovery complete: found {len(files_found)} files "
            f"({total_size / (1024*1024):.2f} MB) in {duration:.2f}s"
        )

        return DiscoveryResult(
            files_found=files_found,
            total_count=len(files_found),
            total_size=total_size,
            errors=errors,
            discovery_duration=duration
        )

    def validate_file_accessibility(self, file_path: Path) -> bool:
        """
        Check if an Excel file can be read and processed

        Args:
            file_path: Path to the Excel file

        Returns:
            True if file is accessible and valid, False otherwise
        """
        try:
            # Check file existence
            if not file_path.exists():
                return False

            # Check if it's a file
            if not file_path.is_file():
                return False

            # Check read permission
            if not os.access(file_path, os.R_OK):
                return False

            # Try to open with openpyxl to validate format
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                workbook.close()
                return True
            except Exception as e:
                self.logger.debug(f"File validation failed for {file_path.name}: {str(e)}")
                return False

        except Exception as e:
            self.logger.error(f"Error validating file {file_path}: {str(e)}")
            return False

    def get_file_metadata(self, file_path: Path) -> ExcelFileInfo:
        """
        Extract metadata from an Excel file

        Args:
            file_path: Path to the Excel file

        Returns:
            ExcelFileInfo with file metadata

        Raises:
            FileNotFoundError: If file doesn't exist
            PermissionError: If file is not readable
        """
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if not os.access(file_path, os.R_OK):
            raise PermissionError(f"File is not readable: {file_path}")

        try:
            # Get file stats
            stat_info = file_path.stat()

            # Calculate file hash
            file_hash = self._calculate_file_hash(file_path)

            # Get sheet count
            sheet_count = 0
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheet_count = len(workbook.sheetnames)
                workbook.close()
            except Exception as e:
                self.logger.warning(f"Could not read sheet count for {file_path.name}: {str(e)}")
                sheet_count = 0

            # Create ExcelFileInfo
            file_info = ExcelFileInfo(
                file_path=file_path,
                file_name=file_path.name,
                file_size=stat_info.st_size,
                file_hash=file_hash,
                created_at=datetime.fromtimestamp(stat_info.st_ctime).isoformat(),
                modified_at=datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                sheet_count=sheet_count
            )

            return file_info

        except (FileNotFoundError, PermissionError):
            raise
        except Exception as e:
            self.logger.error(f"Error extracting metadata from {file_path}: {str(e)}")
            raise

    def watch_directory_for_changes(
        self,
        directory_path: Path,
        callback: callable
    ) -> None:
        """
        Monitor directory for new Excel files (for incremental migration)

        Args:
            directory_path: Directory to monitor
            callback: Function to call when new files are detected

        Raises:
            ValueError: If directory_path is invalid
        """
        # Validate directory
        if not directory_path or not isinstance(directory_path, Path):
            raise ValueError("directory_path must be a valid Path object")

        if not directory_path.exists():
            raise ValueError(f"Directory does not exist: {directory_path}")

        if not directory_path.is_dir():
            raise ValueError(f"Path is not a directory: {directory_path}")

        # This is a placeholder implementation
        # In a production system, this would use file system watchers
        # like watchdog library to monitor for changes
        self.logger.warning(
            "Directory watching is not fully implemented. "
            "Consider using watchdog library for production use."
        )

        # Raise NotImplementedError for now as indicated in contract tests
        raise NotImplementedError(
            "Directory watching requires watchdog library. "
            "This is a placeholder implementation."
        )

    def _calculate_file_hash(self, file_path: Path, algorithm: str = 'md5') -> str:
        """
        Calculate hash of a file

        Args:
            file_path: Path to the file
            algorithm: Hash algorithm to use (md5, sha256, etc.)

        Returns:
            Hex digest of file hash
        """
        hash_obj = hashlib.new(algorithm)

        with open(file_path, 'rb') as f:
            # Read file in chunks to handle large files
            chunk_size = 8192
            while chunk := f.read(chunk_size):
                hash_obj.update(chunk)

        return hash_obj.hexdigest()
